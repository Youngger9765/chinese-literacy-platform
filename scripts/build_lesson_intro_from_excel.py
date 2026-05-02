#!/usr/bin/env python3
"""
Build lesson_intro fallback map from 自學教材清單.xlsx (Issue #1443).

Reads:
  - Sheet "總表": 原課次(col1), 適用年級(col2), 課名(col3), 單元議題(col8), 第六大題標題(col9)
  - Sheet "文言文": same but col 6=單元議題, col 9=第六大題標題

For each row that has a valid lesson code (grade + original lesson number),
builds a lesson_intro dict:
  {
    source: "excel",
    text: "本課探討「{單元議題}」主題，學習「{第六大題標題}」的閱讀方法。",
    unit_topic: "{單元議題}",
    strategy_title: "{第六大題標題}",
  }

Output: /tmp/excel_intro_map.json — dict[lesson_code → intro_dict]

Usage:
    python3 scripts/build_lesson_intro_from_excel.py
    python3 scripts/build_lesson_intro_from_excel.py --output /path/to/map.json
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path

import openpyxl

XLSX_PATH = Path(
    "/Users/young/project/chinese-literacy-platform"
    "/private/curriculum-source/2026-05-01"
    "/1.L1-158新版完成學習單1150415/自學教材清單.xlsx"
)
DEFAULT_OUTPUT = Path("/tmp/excel_intro_map.json")

# Grade label → G-prefix mapping
GRADE_TO_G: dict[int, str] = {
    4: "G4",
    5: "G5",
    6: "G6",
    7: "G7",
    8: "G8",
    9: "G9",
}


def build_lesson_code(grade: int, lesson_num: int) -> str | None:
    """Convert grade + lesson number to lesson code like G4-L1."""
    g = GRADE_TO_G.get(grade)
    if g is None:
        return None
    return f"{g}-L{lesson_num}"


def build_intro_dict(unit_topic: str, strategy_title: str) -> dict:
    """Construct the lesson_intro dict from excel columns."""
    unit_topic = str(unit_topic).strip() if unit_topic else ""
    strategy_title = str(strategy_title).strip() if strategy_title else ""

    parts = []
    if unit_topic:
        parts.append(f"本課探討「{unit_topic}」主題")
    if strategy_title:
        parts.append(f"學習「{strategy_title}」的閱讀方法")

    if not parts:
        return {}

    text = "，".join(parts) + "。"

    result: dict = {"source": "excel", "text": text}
    if unit_topic:
        result["unit_topic"] = unit_topic
    if strategy_title:
        result["strategy_title"] = strategy_title
    return result


def process_sheet(ws, grade_col: int, lesson_col: int, topic_col: int, strategy_col: int,
                  intro_map: dict, sheet_name: str) -> int:
    """Process one worksheet and populate intro_map. Returns count of rows added."""
    added = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Skip header row
            continue
        grade_val = row[grade_col] if len(row) > grade_col else None
        lesson_val = row[lesson_col] if len(row) > lesson_col else None
        topic_val = row[topic_col] if len(row) > topic_col else None
        strategy_val = row[strategy_col] if len(row) > strategy_col else None

        # Skip rows with no grade or lesson number
        if not grade_val or not lesson_val:
            continue
        try:
            grade = int(grade_val)
            lesson_num = int(lesson_val)
        except (TypeError, ValueError):
            continue

        code = build_lesson_code(grade, lesson_num)
        if not code:
            continue

        topic_str = str(topic_val).strip() if topic_val else ""
        strategy_str = str(strategy_val).strip() if strategy_val else ""

        # Skip rows with no useful content
        if not topic_str and not strategy_str:
            continue

        intro = build_intro_dict(topic_str, strategy_str)
        if intro:
            intro_map[code] = intro
            added += 1
    return added


def build_map(xlsx_path: Path = XLSX_PATH) -> dict:
    """Load xlsx and return lesson_code → intro_dict map."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    intro_map: dict = {}

    # === 總表 sheet ===
    # Columns (0-indexed): 0=新課次, 1=原課次, 2=適用年級, 3=課名,
    #   4=文體, 5=類型, 6=目標策略, 7=研發夥伴用, 8=單元議題, 9=第六大題標題
    if "總表" in wb.sheetnames:
        ws_main = wb["總表"]
        added = process_sheet(
            ws_main,
            grade_col=2,    # 適用年級
            lesson_col=1,   # 原課次
            topic_col=8,    # 單元議題
            strategy_col=9, # 第六大題標題
            intro_map=intro_map,
            sheet_name="總表",
        )
        print(f"  總表: {added} lessons added")
    else:
        print("  WARNING: 總表 sheet not found")

    # === 文言文 sheet ===
    # Columns (0-indexed): 0=新課次, 1=原課次, 2=適用年級, 3=課名,
    #   4=文體, 5=類型, 6=單元議題, 7=研發夥伴用, 8=目標策略, 9=第六大題標題
    # NOTE: 文言文 codes map to 文-Ln
    # But we need to distinguish 文言文 from regular grades —
    # check the 適用年級 column to see if it has grade 8 or similar,
    # or rely on 原課次 starting from 1 for 文言文.
    # For now, use build_lesson_code which returns None for non-4..9 grades.
    # 文言文 typically has grade 8 or 9 so they'll map to G8/G9 codes.
    if "文言文" in wb.sheetnames:
        ws_wenyan = wb["文言文"]
        # 文言文 sheet col structure verified: 0=新課次,1=原課次,2=適用年級,
        # 3=課名,4=文體,5=類型,6=單元議題,7=研發夥伴用,8=目標策略,9=第六大題標題
        added = process_sheet(
            ws_wenyan,
            grade_col=2,    # 適用年級
            lesson_col=1,   # 原課次
            topic_col=6,    # 單元議題 (col 6 for 文言文 — one column earlier)
            strategy_col=9, # 第六大題標題
            intro_map=intro_map,
            sheet_name="文言文",
        )
        print(f"  文言文: {added} lessons added")
    else:
        print("  WARNING: 文言文 sheet not found")

    wb.close()
    return intro_map


def main():
    ap = argparse.ArgumentParser(description="Build lesson_intro map from Excel")
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output JSON path")
    ap.add_argument("--xlsx", default=str(XLSX_PATH), help="Path to 自學教材清單.xlsx")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    print(f"Reading: {xlsx_path}")
    intro_map = build_map(xlsx_path)
    print(f"Total: {len(intro_map)} lessons mapped")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(intro_map, f, ensure_ascii=False, indent=2, sort_keys=True)

    print(f"Saved: {output_path}")

    # Show a few samples
    samples = list(intro_map.items())[:5]
    print("\nSamples:")
    for code, intro in samples:
        print(f"  {code}: {intro['text'][:60]}...")


if __name__ == "__main__":
    main()
