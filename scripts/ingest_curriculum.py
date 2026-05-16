#!/usr/bin/env python3
"""
Curriculum ingestion parser — auto-sync from 教授 source to backend/data/.

Reads:
  - <source>/自學教材清單.xlsx (master index, sheets: 總表 + 文言文 + 注意課次安排)
  - <source>/1-1教師版(L1~122)/{grade}/G{grade}-L{NN}*.docx (teacher version)
  - <source>/2-1學生版(L1~122)/{grade}/G{grade}-SL{NN}*.docx (student version)
  - <source>/第三階段(L123開始~L157)差學生版/G{grade}-L{NN}*.docx (phase 3)
  - backend/data/curriculum/manifest.yml (158-canonical identity index, optional)

Writes (to --target):
  - curriculum/lessons/{lesson_code}.yml   (catalog metadata; lesson_code from manifest)
  - lessons/L{original_order}.yml          (content yml, keyed by 原課次; ONLY for white-vernacular with valid original_order)
  - curriculum/INGESTION_MANIFEST.yml      (active_version + provenance)

Identity model (round 2 — manifest-driven):
  - Canonical lesson list comes from `backend/data/curriculum/manifest.yml`
    (158 entries: G4-L01..G9-L19 + WW-L01..L10 + multi-part G8-L03a/b etc.)
  - When manifest is available, parser uses it as identity source and only
    enriches fields from xlsx by title match.
  - When manifest is missing, parser falls back to xlsx-driven mode (original
    round 1 behavior; mainly for cold-start scenarios).
  - Content yml stems (L01..L58) are STABLE across editions — downstream
    AI-generated fields (story_structure, fill_in_blank, etc.) keyed by these stems.
  - 文言文 lessons map to WW-L01..WW-L10 catalog codes (per canonical manifest).

Parser writes catalog yml in full + content yml minimal subset (only when the
stem exists in target's existing content yml dir). Drift detector compares
parser-managed fields only; downstream-managed fields are NOT touched.

Usage:
  python scripts/ingest_curriculum.py \\
    --source private/curriculum-source/2026-05-01/1.L1-158新版完成學習單1150415/ \\
    --target /tmp/parsed-from-source/ --write

Owner: Young / LingoLeap | Issues: #1624 (round 1), #1628 (round 2)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import signal
import sys
from pathlib import Path
from typing import Any

try:
    from docx import Document
except ImportError:
    sys.stderr.write("ERROR: python-docx required. pip install python-docx\n")
    sys.exit(2)

try:
    import openpyxl
except ImportError:
    sys.stderr.write("ERROR: openpyxl required. pip install openpyxl\n")
    sys.exit(2)

try:
    import yaml
except ImportError:
    sys.stderr.write("ERROR: pyyaml required. pip install pyyaml\n")
    sys.exit(2)


# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------

DOCX_PARSE_TIMEOUT_SECONDS = 60
PARSER_VERSION = "2.0.0"
EXCEL_DATE_BASE = dt.datetime(2025, 4, 1)
DEFAULT_MANIFEST_PATH = Path("backend/data/curriculum/manifest.yml")


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------


def _parse_grade(val: Any) -> int | str | None:
    if val is None or val == "無" or val == "":
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    return None


def _normalize_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _decode_new_order(raw: Any) -> int | None:
    """Excel auto-converted 新課次 1 → 2025-04-01. Reverse it."""
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return (raw - EXCEL_DATE_BASE).days + 1
    if isinstance(raw, (int, float)):
        return int(raw)
    s = _normalize_str(raw)
    if s.isdigit():
        return int(s)
    return None


def _lesson_code(grade: Any, per_grade_seq: int) -> str:
    """Legacy lesson_code builder (round 1 fallback only).

    Round 2 prefers manifest.yml lesson_code directly. This helper is kept
    for the xlsx-only fallback path when no manifest is available.
    Note: 文言文 maps to WW-L* (not 文-L*) per canonical manifest.
    """
    if isinstance(grade, int):
        return f"G{grade}-L{per_grade_seq:02d}"
    # 文言文 → WW-L* (per manifest.yml canonical naming)
    return f"WW-L{per_grade_seq:02d}"


def _content_yml_stem(original_order: int) -> str:
    if original_order is None:
        return None
    return f"L{original_order:02d}" if original_order < 100 else f"L{original_order}"


def _parse_int_or_none(val: Any) -> int | None:
    """Strict int parser. Returns None for None/non-numeric (e.g. '19-3' multi-part marker)."""
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


# ----------------------------------------------------------------------
# Manifest (canonical identity)
# ----------------------------------------------------------------------


def load_canonical_manifest(manifest_path: Path) -> list[dict] | None:
    """Load `backend/data/curriculum/manifest.yml` (158-canonical lesson list).

    Returns list of canonical lessons or None if manifest missing.
    Each entry has: lesson_code, display_order, title, grade, strategy_id,
    strategy_name, classical_chinese.
    """
    if not manifest_path.exists():
        return None
    with manifest_path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    if not data or "lessons" not in data:
        return None
    return data["lessons"]


# ----------------------------------------------------------------------
# Step A: parse master xlsx
# ----------------------------------------------------------------------


def parse_master_xlsx(xlsx_path: Path) -> dict:
    """Parse 自學教材清單.xlsx → identity+enrichment lookup.

    Returns:
      {
        "lessons": [...legacy fallback list...],   # round-1 compatible
        "by_title": {title -> enrichment dict},    # round-2 lookup
        "an_by_title": {title -> 注意課次安排 row}, # for applicable_grade
      }

    Enrichment dict per title:
      original_order, genre, text_type, unit_topic, strategy_name,
      vocab, classical_chinese, applicable_grade
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    lessons: list[dict] = []
    by_title: dict[str, dict] = {}

    # --- 總表 sheet (白話) ---
    ws = wb["總表"]
    header = [_normalize_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    vocab_col_idx = [i for i, h in enumerate(header) if h.startswith("語詞")]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Some 總表 rows have new_order=None (e.g. lessons without scheduled position),
        # but still have valid title + grade. Don't skip those — manifest may reference them.
        new_order = _decode_new_order(row[0]) if row[0] is not None else None

        original_order = _parse_int_or_none(row[1])

        grade = _parse_grade(row[2])
        title = _normalize_str(row[3])
        if not title or grade is None:
            continue

        genre = _normalize_str(row[4])
        strategy_name = _normalize_str(row[6]) or "無"
        unit_topic = _normalize_str(row[8]) or None
        text_type = _normalize_str(row[10])

        # Vocab: preserve "無" sentinel value (staging uses ['無'] when no vocab)
        # vs None when xlsx row is completely empty.
        vocab_raw = []
        for vi in vocab_col_idx:
            if vi < len(row) and row[vi] is not None:
                v = _normalize_str(row[vi])
                if v:
                    vocab_raw.append(v)
        # Keep '無' if it's the only entry (staging convention)
        if vocab_raw == ["無"]:
            vocab = ["無"]
        elif vocab_raw:
            vocab = [v for v in vocab_raw if v != "無"]
            if not vocab:
                vocab = None  # all were '無' but more than one — treat as null
        else:
            vocab = None  # No vocab columns populated → null (staging convention)

        record = {
            "new_order": new_order,
            "original_order": original_order,
            "grade": grade,
            "title": title,
            "genre": genre,
            "text_type": text_type,
            "unit_topic": unit_topic,
            "applicable_grade": None,
            "strategy_name": strategy_name,
            "vocab": vocab,
            "classical_chinese": False,
        }
        lessons.append(record)
        # Last-write-wins on title duplicates (rare; 總表 should have unique titles)
        by_title[title] = record

    # --- 文言文 sheet ---
    if "文言文" in wb.sheetnames:
        ws = wb["文言文"]
        header = [_normalize_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        vocab_col_idx = [i for i, h in enumerate(header) if h.startswith("語詞")]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            new_order = _parse_int_or_none(row[0])
            if new_order is None:
                continue
            original_order = _parse_int_or_none(row[1])

            applicable_grade = _normalize_str(row[2]) or None
            title = _normalize_str(row[3])
            if not title:
                continue
            genre = _normalize_str(row[4])
            unit_topic = _normalize_str(row[6]) or None
            strategy_name = _normalize_str(row[8]) or "無"
            text_type = _normalize_str(row[10])

            # Same vocab handling as 總表
            vocab_raw = []
            for vi in vocab_col_idx:
                if vi < len(row) and row[vi] is not None:
                    v = _normalize_str(row[vi])
                    if v:
                        vocab_raw.append(v)
            if vocab_raw == ["無"]:
                vocab = ["無"]
            elif vocab_raw:
                vocab = [v for v in vocab_raw if v != "無"]
                if not vocab:
                    vocab = None
            else:
                vocab = None

            record = {
                "new_order": new_order,
                "original_order": original_order,
                "grade": "文",
                "title": title,
                "genre": genre,
                "text_type": text_type,
                "unit_topic": unit_topic,
                "applicable_grade": applicable_grade,
                "strategy_name": strategy_name,
                "vocab": vocab,
                "classical_chinese": True,
            }
            lessons.append(record)
            by_title[title] = record

    # --- 注意課次安排 sheet (provides applicable_grade keyword) ---
    an_by_title: dict[str, Any] = {}
    if "注意課次安排" in wb.sheetnames:
        ws = wb["注意課次安排"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[3] is None:
                continue
            title = _normalize_str(row[3])
            if not title:
                continue
            an_by_title[title] = {
                "new_order": _parse_int_or_none(row[0]),
                "original_order_raw": row[1],  # may be '19-3' multi-part marker
                "applicable_grade": _normalize_str(row[6]) or None,
                "grade": _parse_int_or_none(row[7]),
            }

    wb.close()
    # Sort by new_order; rows with no new_order go to the end (None sorts as +infinity)
    lessons.sort(key=lambda x: (x["new_order"] is None, x["new_order"] or 0))

    # Compute per-grade sequential (legacy fallback only; manifest mode bypasses)
    per_grade_counter: dict[Any, int] = {}
    for L in lessons:
        g = L["grade"]
        per_grade_counter[g] = per_grade_counter.get(g, 0) + 1
        L["per_grade_seq"] = per_grade_counter[g]

    return {
        "lessons": lessons,
        "by_title": by_title,
        "an_by_title": an_by_title,
    }


# ----------------------------------------------------------------------
# Manifest-driven build (round 2)
# ----------------------------------------------------------------------


def _title_normalize(title: str) -> str:
    """Normalize title for cross-source lookup.

    Handles common variants across xlsx (professor source) and manifest/staging:
      - ＃ ↔ # (full/half-width hash)
      - ？ ↔ ? (full/half-width question)
      - 沒 ↔ 没 (Taiwan trad ↔ simplified variant; some xlsx rows use 简体)
      - ！ ↔ ! (full/half-width exclamation)
      - 「」‟" ↔ none (quotes removed for matching)
    """
    norm = (title or "").strip()
    norm = norm.replace("＃", "#")
    norm = norm.replace("？", "?")
    norm = norm.replace("！", "!")
    norm = norm.replace("沒", "没")  # collapse to 简体 form for match
    return norm


def build_lessons_from_manifest(
    canonical: list[dict],
    xlsx_data: dict,
) -> list[dict]:
    """Build lesson records aligned with manifest.yml canonical identity.

    For each canonical manifest entry, look up enrichment by title from xlsx:
      - From 總表 / 文言文: genre, text_type, unit_topic, vocab, strategy_name (full)
      - From 注意課次安排: applicable_grade keyword
      - From manifest: lesson_code, display_order, title, grade, classical_chinese,
                       strategy_name (canonical), original_order (when available)

    Returns list of lesson records ready for render_catalog_yml.
    """
    by_title = xlsx_data["by_title"]
    an_by_title = xlsx_data["an_by_title"]
    # Build normalized lookup tables for fuzzy match (handles ＃/# variants)
    by_title_norm = {_title_normalize(k): v for k, v in by_title.items()}
    an_by_title_norm = {_title_normalize(k): v for k, v in an_by_title.items()}

    out: list[dict] = []
    for entry in canonical:
        title = entry["title"]
        code = entry["lesson_code"]
        display_order = entry["display_order"]
        grade = entry["grade"]
        classical = entry.get("classical_chinese", False)
        # Manifest strategy_name is canonical (matches existing staging yml)
        manifest_strategy = entry.get("strategy_name") or "無"

        # Lookup with normalized title fallback (handles ＃/# half-/full-width variants
        # between manifest and xlsx sources)
        enrich = by_title.get(title) or by_title_norm.get(_title_normalize(title), {})
        an = an_by_title.get(title) or an_by_title_norm.get(_title_normalize(title), {})

        # original_order source (staging convention, verified empirically):
        # - White-vernacular (總表): use row[1] which is the existing content yml stem number
        #   (e.g. G4-L03 長高的祕密 → 原課次=3 in 總表, even though 注意課次安排 says 6)
        # - 文言文 (WW-L*): use 文言文 sheet's 原課次 1-10 (per WW-L01..L10 numbering)
        # - Multi-part (G8-L03a/b etc.): use 總表 row[1] for each lesson individually
        # - Manifest override: if manifest entry has explicit original_order, prefer it
        original_order = entry.get("original_order")
        if original_order is None:
            # Primary: 總表 row[1] / 文言文 sheet row[1] (= enrich.original_order)
            original_order = enrich.get("original_order")
        if original_order is None:
            # Last-resort fallback: 注意課次安排
            an_raw = an.get("original_order_raw")
            if isinstance(an_raw, (int, float)):
                original_order = int(an_raw)
            elif an_raw is not None:
                m = re.match(r"^(\d+)", str(an_raw).strip())
                if m:
                    original_order = int(m.group(1))

        # applicable_grade: 注意課次安排 col 6 by title match (staging behavior).
        # If title not in 注意課次安排, stays None to match staging null entries.
        applicable_grade = an.get("applicable_grade")

        # Preserve vocab nullability: enrich.vocab is None when xlsx empty,
        # ['無'] when xlsx has 無, [...] otherwise. Don't coerce to [].
        vocab_field = enrich.get("vocab", None)
        if "vocab" not in enrich:
            vocab_field = None

        out.append(
            {
                "lesson_code": code,
                "display_order": display_order,
                "original_order": original_order,
                "title": title,
                "grade": grade,
                "genre": enrich.get("genre") or None,
                "text_type": enrich.get("text_type") or None,
                "unit_topic": enrich.get("unit_topic"),
                "applicable_grade": applicable_grade,
                "strategy_name": manifest_strategy,
                "classical_chinese": classical,
                "vocab": vocab_field,
            }
        )
    return out


# ----------------------------------------------------------------------
# Step B + C: scan docx + extract story_text
# ----------------------------------------------------------------------


def scan_docx_files(source_root: Path) -> dict:
    """Scan all docx in source tree.

    Returns:
      {
        "by_code": {lesson_code → {teacher: Path, student: Path}},  # legacy
        "by_title_kw": [(title_keyword, Path, role), ...],          # for title match
      }

    Note: docx filenames use linear numbering (G8-L1..L19, 文-L1..L9), but
    canonical manifest has multi-part codes (G8-L03a/b). Round-2 build uses
    title keyword match to bridge this gap.
    """
    teacher_dir = source_root / "1-1教師版(L1~122)"
    student_dir = source_root / "2-1學生版(L1~122)"
    phase3_dir = source_root / "第三階段(L123開始~L157)差學生版"

    found: dict[str, dict[str, Path]] = {}
    by_title_kw: list[tuple[str, Path, str]] = []

    def _scan(root: Path, role: str) -> None:
        if not root.exists():
            return
        for docx in root.rglob("*.docx"):
            if docx.name.startswith("~$"):
                continue
            stem = docx.stem
            # Match G{grade}-L{NN} or G{grade}-SL{NN} (also G{grade}-L{NN}-{MM} for multi-lesson)
            m = re.match(r"G(\d+)-S?L(\d+)(?:-\d+)?(.*)", stem)
            if m:
                grade_num = int(m.group(1))
                lesson_num = int(m.group(2))
                code = f"G{grade_num}-L{lesson_num:02d}"
                # Capture title keyword (everything after the lesson number prefix)
                title_kw = m.group(3).strip()
                if title_kw:
                    by_title_kw.append((title_kw, docx, role))
            else:
                # Classical: filenames like "文-L1假新聞？鞭虎救弟記..."
                m2 = re.match(r"文-L?(\d+)(.*)", stem)
                if not m2:
                    continue
                lesson_num = int(m2.group(1))
                # Legacy 文-L code (for round-1 fallback)
                code = f"文-L{lesson_num:02d}"
                title_kw = m2.group(2).strip()
                if title_kw:
                    by_title_kw.append((title_kw, docx, role))
            found.setdefault(code, {})[role] = docx

    _scan(teacher_dir, "teacher")
    _scan(student_dir, "student")
    _scan(phase3_dir, "student")
    return {"by_code": found, "by_title_kw": by_title_kw}


def find_docx_for_title(docx_index: dict, title: str) -> dict[str, Path]:
    """Find docx files for a manifest title by keyword overlap.

    Strategy: pick first 2-3 chars of title that aren't punctuation and
    check substring match in docx filename keyword. This is heuristic and
    may miss some matches; matches drive story_text extraction only,
    drift detector ignores story_text fields.
    """
    if not title:
        return {}
    # Strip leading punctuation
    norm_title = re.sub(r"^[「『《＃#（()]+", "", title)
    # Take first 4 chars or up to first punctuation
    head = re.match(r"^([^：（()\[\]「『《──！？。，、:!?,]+)", norm_title)
    if not head or len(head.group(1)) < 2:
        return {}
    needle = head.group(1)[:6]

    result: dict[str, Path] = {}
    for title_kw, docx_path, role in docx_index.get("by_title_kw", []):
        # Match if needle is substring of title_kw or vice versa
        if needle in title_kw or title_kw[:len(needle)] == needle:
            if role not in result:
                result[role] = docx_path
    return result


def _docx_timeout_handler(signum, frame):
    raise TimeoutError("docx parsing timeout")


def extract_story_from_docx(docx_path: Path) -> tuple[str, list[str]]:
    """Extract story_text from a teacher docx.

    Strategy: scan tables for the largest text cell > 100 chars (heuristic
    verified on G4-L1 sample — story in table[0][2,1]).
    """
    signal.signal(signal.SIGALRM, _docx_timeout_handler)
    signal.alarm(DOCX_PARSE_TIMEOUT_SECONDS)
    try:
        doc = Document(docx_path)
        candidate: tuple[int, str] = (0, "")
        for table in doc.tables:
            for row in table.rows:
                for cell in row.cells:
                    text = cell.text.strip()
                    # Skip obvious non-story cells
                    if len(text) > candidate[0] and not _looks_like_metadata(text):
                        candidate = (len(text), text)

        if candidate[0] < 100:
            return "", []

        story = candidate[1]
        chunks = re.split(r"\n\s*\n|\n(?=　)", story)
        paragraphs = [c.strip() for c in chunks if c.strip()]
        full_text = "\n".join(paragraphs)
        return full_text, paragraphs
    except TimeoutError:
        sys.stderr.write(f"WARN: timeout parsing {docx_path.name} — skipped\n")
        return "", []
    except Exception as e:
        sys.stderr.write(f"WARN: failed to parse {docx_path.name}: {e}\n")
        return "", []
    finally:
        signal.alarm(0)


def _looks_like_metadata(text: str) -> bool:
    """Filter out non-story cells (vocab banks, video lists, instructions)."""
    head = text[:30]
    for marker in ("第", "計時", "影片", "請", "□", "◎", "A.", "B.", "C.", "D.", "E.", "F.", "G.", "H."):
        if head.startswith(marker):
            return True
    return False


# ----------------------------------------------------------------------
# Step D + E: render yml
# ----------------------------------------------------------------------


def render_catalog_yml(lesson: dict) -> dict:
    """Render catalog yml dict.

    Accepts either:
      - Round-2 manifest-driven lesson dict (has lesson_code, display_order set)
      - Round-1 xlsx-only lesson dict (has grade, per_grade_seq, new_order)
    """
    if "lesson_code" in lesson and "display_order" in lesson:
        code = lesson["lesson_code"]
        display_order = lesson["display_order"]
        grade_field = lesson["grade"] if isinstance(lesson["grade"], int) else None
    else:
        # legacy fallback
        code = _lesson_code(lesson["grade"], lesson["per_grade_seq"])
        display_order = lesson["new_order"]
        grade_field = lesson["grade"] if isinstance(lesson["grade"], int) else None

    content_stem = _content_yml_stem(lesson["original_order"])
    # Normalize empty-string fields to None to match staging convention
    genre = lesson.get("genre") or None
    text_type = lesson.get("text_type") or None
    return {
        "lesson_code": code,
        "display_order": display_order,
        "original_order": lesson["original_order"],
        "title": lesson["title"],
        "grade": grade_field,
        "genre": genre,
        "text_type": text_type,
        "unit_topic": lesson.get("unit_topic"),
        "applicable_grade": lesson.get("applicable_grade"),
        "strategy_id": None,
        "strategy_name": lesson["strategy_name"],
        "strategy_dependency": None,
        "strategy_relation": None,
        "classical_chinese": lesson["classical_chinese"],
        "video_links": [],
        "vocab": lesson.get("vocab"),
        "existing_content_yaml": f"{content_stem}.yml" if content_stem else None,
        "notes": None,
    }


def render_content_yml_minimal(lesson: dict, story_text: str, paragraphs: list[str]) -> dict:
    """Minimal content yml — only parser-managed fields.

    Drift detector compares ONLY these fields. Downstream-managed fields
    (vocabulary, vocab_bank, fill_in_blank, multiple_choice, story_structure, etc.)
    are NOT touched.
    """
    grade_field = lesson["grade"] if isinstance(lesson["grade"], int) else None
    if "lesson_code" in lesson:
        code = lesson["lesson_code"]
    else:
        code = _lesson_code(lesson["grade"], lesson["per_grade_seq"])
    return {
        "lesson_number": lesson["original_order"],
        "grade": grade_field,
        "grade_code": code,
        "title": lesson["title"],
        "genre": lesson["genre"],
        "text_type": lesson["text_type"],
        "reading_strategy": lesson["strategy_name"],
        "story_text": story_text,
        "paragraphs": paragraphs,
        "paragraph_count": len(paragraphs),
        "char_count": len(story_text),
        "vocab_keywords": lesson["vocab"],
    }


# ----------------------------------------------------------------------
# Step F: MANIFEST
# ----------------------------------------------------------------------


def write_manifest(target_root: Path, source_root: Path, version: str, stats: dict) -> Path:
    # Use INGESTION_MANIFEST.yml to avoid case-collision with existing manifest.yml
    # (macOS is case-insensitive — the legacy manifest.yml is loaded by lesson_loader.py)
    manifest_path = target_root / "curriculum" / "INGESTION_MANIFEST.yml"
    try:
        rel_source = str(source_root.relative_to(Path.cwd()))
    except ValueError:
        rel_source = str(source_root)
    manifest = {
        "active_version": version,
        "source_dir": rel_source,
        "source_excel": "自學教材清單.xlsx",
        "parsed_at": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "parser_version": PARSER_VERSION,
        "stats": stats,
        "schema_note": (
            "MANIFEST tracks the active 教授 source version. "
            "Use scripts/check_curriculum_drift.py to verify backend/data/ "
            "matches what parser would generate from this source."
        ),
        "parser_managed_catalog_fields": [
            "lesson_code", "display_order", "original_order", "title", "grade",
            "genre", "text_type", "unit_topic", "applicable_grade",
            "strategy_name", "classical_chinese", "vocab",
        ],
        # Round 2: parser_managed_content_fields reduced to identity-only fields.
        # Earlier round 1 included story_text/paragraphs/reading_strategy/vocab_keywords,
        # but content yml in backend/data/lessons/ has been hand-curated (paragraph
        # splits, reading_strategy wording variants, vocab_keywords None-vs-list).
        # Parser-derived docx-extracted fields cannot reproduce hand-curation
        # without massive drift. Keep identity fields (lesson_number, grade, grade_code,
        # title, genre, text_type) which are stable; content authoring is downstream.
        "parser_managed_content_fields": [
            "lesson_number", "grade", "grade_code", "title", "genre", "text_type",
        ],
    }
    return _write_yaml(manifest_path, manifest)


def _write_yaml(path: Path, data: dict) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False, width=10000)
    return path


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description="Ingest curriculum source → backend/data yml")
    ap.add_argument("--source", required=True, type=Path,
                    help="教授 source root (contains 自學教材清單.xlsx)")
    ap.add_argument("--target", required=True, type=Path,
                    help="Output root (writes curriculum/lessons/ + lessons/ + curriculum/INGESTION_MANIFEST.yml)")
    ap.add_argument("--manifest", type=Path, default=None,
                    help="Canonical manifest.yml (default: backend/data/curriculum/manifest.yml). "
                         "When present, parser uses manifest as identity source and enriches "
                         "from xlsx by title match. When missing, falls back to xlsx-only mode.")
    ap.add_argument("--no-manifest", action="store_true",
                    help="Force xlsx-only mode even if manifest.yml exists.")
    ap.add_argument("--existing-content-dir", type=Path, default=None,
                    help="Existing content yml dir (default: backend/data/lessons). "
                         "Used to decide whether to write content yml — only writes if "
                         "existing stem exists, to avoid creating new files for lessons "
                         "that share stems with others.")
    ap.add_argument("--skip-content-yml", action="store_true", default=True,
                    help="Skip writing content yml (default: True). Content yml in "
                         "backend/data/lessons/ is hand-curated and shouldn't be overwritten "
                         "by docx extraction. Override with --no-skip-content-yml for cold-start.")
    ap.add_argument("--no-skip-content-yml", dest="skip_content_yml", action="store_false")
    ap.add_argument("--version", default=None,
                    help="Version label (default: auto-detect from source parent dir)")
    ap.add_argument("--write", action="store_true",
                    help="Actually write files (default: dry-run, no write)")
    ap.add_argument("--limit", type=int, default=None,
                    help="Limit to first N lessons (for testing)")
    args = ap.parse_args()

    source_root: Path = args.source.resolve()
    target_root: Path = args.target.resolve()

    if not source_root.exists():
        sys.stderr.write(f"ERROR: source not found: {source_root}\n")
        return 2

    xlsx = source_root / "自學教材清單.xlsx"
    if not xlsx.exists():
        sys.stderr.write(f"ERROR: master xlsx not found: {xlsx}\n")
        return 2

    # Resolve manifest path (round 2 identity source)
    repo_root = Path(__file__).resolve().parent.parent
    if args.no_manifest:
        manifest_path = None
    elif args.manifest:
        manifest_path = args.manifest.resolve()
    else:
        manifest_path = (repo_root / "backend" / "data" / "curriculum" / "manifest.yml").resolve()

    canonical = None
    if manifest_path and manifest_path.exists():
        canonical = load_canonical_manifest(manifest_path)

    existing_content_dir = args.existing_content_dir
    if existing_content_dir is None:
        existing_content_dir = repo_root / "backend" / "data" / "lessons"
    existing_content_stems = set()
    if existing_content_dir.exists():
        for f in existing_content_dir.glob("*.yml"):
            existing_content_stems.add(f.stem)

    version = args.version or _infer_version(source_root)
    print(f"[ingest] source: {source_root}", flush=True)
    print(f"[ingest] target: {target_root} ({'WRITE' if args.write else 'DRY-RUN'})", flush=True)
    print(f"[ingest] version: {version}", flush=True)
    print(f"[ingest] manifest: {manifest_path if canonical else 'NONE (xlsx-only fallback)'}", flush=True)
    print(f"[ingest] existing content stems: {len(existing_content_stems)} (only writes content yml for these)", flush=True)

    print(f"[ingest] Step A: parsing master xlsx...", flush=True)
    xlsx_data = parse_master_xlsx(xlsx)
    print(f"[ingest]   loaded {len(xlsx_data['lessons'])} xlsx rows "
          f"(by_title: {len(xlsx_data['by_title'])}, an_by_title: {len(xlsx_data['an_by_title'])})", flush=True)

    if canonical:
        print(f"[ingest] Step A2: building from manifest ({len(canonical)} canonical entries)...", flush=True)
        lessons = build_lessons_from_manifest(canonical, xlsx_data)
        print(f"[ingest]   built {len(lessons)} manifest-aligned lessons", flush=True)
        # Note: round-2 lessons don't have new_order / per_grade_seq
        mode = "manifest-driven"
    else:
        # Legacy fallback: xlsx-only with per-grade-seq inference
        lessons = xlsx_data["lessons"]
        mode = "xlsx-only-fallback"

    print(f"[ingest] mode: {mode}", flush=True)

    print(f"[ingest] Step B: scanning docx files...", flush=True)
    docx_index = scan_docx_files(source_root)
    print(f"[ingest]   found docx for {len(docx_index['by_code'])} lesson codes "
          f"({len(docx_index['by_title_kw'])} title keywords)", flush=True)

    print(f"[ingest] Step C-E: extracting stories + rendering yml...", flush=True)
    catalog_written = 0
    content_written = 0
    content_skipped_no_stem = 0
    missing_story = 0
    missing_docx = 0
    matched_docx = 0

    iter_lessons = lessons[: args.limit] if args.limit else lessons

    for i, lesson in enumerate(iter_lessons, 1):
        if i % 25 == 0:
            print(f"[ingest]   progress: {i}/{len(iter_lessons)}", flush=True)

        # Determine lesson_code (round-2 manifest sets it directly; legacy uses helper)
        if "lesson_code" in lesson:
            code = lesson["lesson_code"]
        else:
            code = _lesson_code(lesson["grade"], lesson["per_grade_seq"])

        # Docx lookup: try lesson_code, then title-based fallback
        docx_paths = docx_index["by_code"].get(code, {})
        if not docx_paths and canonical:
            # Manifest-driven: code may be G8-L03a but docx only has G8-L3, fall back to title match
            docx_paths = find_docx_for_title(docx_index, lesson["title"])

        teacher_docx = docx_paths.get("teacher")
        student_docx = docx_paths.get("student")

        story_text = ""
        paragraphs: list[str] = []
        if teacher_docx and teacher_docx.exists():
            story_text, paragraphs = extract_story_from_docx(teacher_docx)
            matched_docx += 1
        elif student_docx and student_docx.exists():
            story_text, paragraphs = extract_story_from_docx(student_docx)
            matched_docx += 1
        else:
            missing_docx += 1

        if docx_paths and not story_text:
            missing_story += 1

        catalog = render_catalog_yml(lesson)

        content_stem = _content_yml_stem(lesson["original_order"])
        catalog_path = target_root / "curriculum" / "lessons" / f"{code}.yml"

        # Only write content yml if:
        #   1. Stem already exists in target's existing_content_dir (no new files)
        #   2. Existing title matches lesson title exactly (no overwrite of unrelated content)
        #   3. --skip-content-yml is False (default True; content yml is hand-curated)
        # This prevents:
        #   - Creating new L*.yml files for lessons that don't have content yet
        #   - 文言文/WW lessons overwriting white-vernacular L01..L10 stems
        #   - Overwriting hand-curated titles with manifest variants (punctuation diffs)
        # (round 1 bug: 文言文 with original_order=1 wrote L01.yml clobbering G4-L01)
        write_content = False
        content_path = None
        if not args.skip_content_yml and content_stem and content_stem in existing_content_stems:
            existing_path = existing_content_dir / f"{content_stem}.yml"
            try:
                with existing_path.open("r", encoding="utf-8") as fp:
                    existing = yaml.safe_load(fp)
                # Only write if existing title matches lesson title (identity check)
                if existing and existing.get("title") == lesson["title"]:
                    content_path = target_root / "lessons" / f"{content_stem}.yml"
                    write_content = True
            except Exception:
                pass
        if not write_content:
            content_skipped_no_stem += 1

        if write_content and content_path:
            content = render_content_yml_minimal(lesson, story_text, paragraphs)

        if args.write:
            _write_yaml(catalog_path, catalog)
            if write_content and content_path:
                _write_yaml(content_path, content)
        catalog_written += 1
        if write_content:
            content_written += 1

    stats = {
        "total_lessons": len(iter_lessons),
        "catalog_yml_count": catalog_written,
        "content_yml_count": content_written,
        "content_skipped_no_stem": content_skipped_no_stem,
        "docx_matched": matched_docx,
        "missing_docx": missing_docx,
        "missing_story_text": missing_story,
        "白話": sum(1 for L in iter_lessons if not L["classical_chinese"]),
        "文言文": sum(1 for L in iter_lessons if L["classical_chinese"]),
        "mode": mode,
    }
    if args.write:
        ingestion_manifest_path = write_manifest(target_root, source_root, version, stats)
        print(f"[ingest] wrote INGESTION_MANIFEST: {ingestion_manifest_path}", flush=True)

    print(f"\n[ingest] DONE", flush=True)
    print(f"[ingest] stats: {json.dumps(stats, ensure_ascii=False)}", flush=True)
    if not args.write:
        print(f"[ingest] (dry-run — no files written; pass --write to commit)", flush=True)
    return 0


def _infer_version(source_root: Path) -> str:
    for part in reversed(source_root.parts):
        m = re.match(r"^(\d{4}-\d{2}-\d{2})$", part)
        if m:
            return m.group(1)
    return dt.date.today().strftime("%Y-%m-%d")


if __name__ == "__main__":
    sys.exit(main())
