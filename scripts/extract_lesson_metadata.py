#!/usr/bin/env python3
"""extract_lesson_metadata.py — 課程簡介 / 文體 / 影片連結, from 自學教材總表.xlsx (#2683).

WHERE THIS CAME FROM
--------------------
I reported 課程簡介 as unfixable — "the DOCX has no such section" — and stopped. That
was true and irrelevant: the first edition never took the intro from the DOCX either.
`scripts/build_lesson_intro_from_excel.py` built it from a spreadsheet, and the second
edition ships the same spreadsheet with more columns than before.

The lesson to carry forward: before declaring a field unobtainable, look at how the
previous edition obtained it. The answer was one `ls scripts/` away.

WHAT THE SPREADSHEET HAS THAT THE WORKSHEET DOES NOT
----------------------------------------------------
    1.總表                 854 rows — 單元議題, 閱讀聚光燈策略, 文體, 類型
    2.體育生的品格聚光燈    the 品格教育 collection, same columns
    3.文言文               the 文言文 collection, same columns
    4.影片連結             actual YouTube URLs per lesson

So this fills four gaps at once: the intro, the genre/category taxonomy that has been
empty since the re-ink, and real links for 知識補給站 — the DOCX only had titles and
durations, which is why that section had no cross-check.

MATCHED ON TITLE, NEVER ON LESSON NUMBER
----------------------------------------
The sheet's 課次 is first-edition numbering — its 「舊課次」 column is literally that.
Joining on it would repeat the mistake that put a bus interior on a sprinting lesson
and another lesson's passage into 重點朗讀. Titles are compared with punctuation and
width folded away, since those drift between editions.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

DEFAULT_XLSX = Path(
    "/Users/young/Downloads/自學教材給數位團隊/自學教材總表.xlsx"
)

_PUNCT = "「」『』（）()，,。.？?！!：:；;、－-—～~　 "

# 文體 → the four categories the API contract allows.
GENRE_TO_CATEGORY = {
    "記敘文": "Fable",
    "說明文": "Science",
    "説明文": "Science",
    "議論文": "History",
    "文言文": "History",
    "應用文": "Daily",
    "抒情文": "Fable",
}


#: Character pairs the two sources spell differently. Not a normalisation any library
#: performs — these are distinct code points, and 沒/没 and 毀/燬 are simply the
#: variant each editor happened to type.
_VARIANTS = str.maketrans({"没": "沒", "燬": "毀", "喫": "吃", "牠": "它", "颱": "台"})


def norm_title(t: str | None) -> str:
    t = unicodedata.normalize("NFKC", str(t or ""))
    # Editing artefacts that lead a title: a stray ordinal (「-18牧羊少年的逆轉勝」), a
    # drafting marker (「~25運動記錄的突破與公平」), a collection prefix
    # (「信任_下一球，我相信你」, 「多文本-未解之謎」).
    t = re.sub(r"^[-－~～#＃\s]*\d{0,3}\s*", "", t)
    t = re.sub(r"^(多文本|信任|勇氣|合作)[-_－—]\s*", "", t)
    return "".join(c for c in t if c not in _PUNCT).translate(_VARIANTS)


#: Pairs a human confirmed, where the two titles differ by more than a threshold can
#: safely bridge. Named individually rather than by loosening the fuzzy cutoff — a
#: lower cutoff would quietly change every other join to recover this one.
ALIASES = {
    # Same lesson, different phrasing of the same four charts.
    "四張圖看地球暖化": "從四張圖看地球暖化的現象",
}

#: Titles that look like near-misses but must NOT be joined:
#:   正太與小豬          — two spreadsheet rows share the opening; picking one is a coin flip
#:   多文本-巨石陣+摩艾石像 — one lesson drawing on two spreadsheet rows, not a 1:1 row


def lookup(meta: dict[str, dict], title: str) -> dict | None:
    """Find a lesson's row, tolerating the ways the two sources spell a title.

    Exact, then a shared opening, then a high-cutoff fuzzy match — each step only
    when it is unambiguous, because a title join that guesses is how the first
    edition's covers ended up on the wrong lessons. Every non-exact pairing is
    recorded in metadata.yml so it can be read back.
    """
    key = norm_title(title)
    key = ALIASES.get(key, key)
    if key in meta:
        return meta[key]

    # A shared opening, where one side carries a subtitle the other drops
    # (「正太與小豬」 / 「正太與小豬：武僧的養成之路」). Prefix only, and only when the
    # shorter side is long enough to identify a lesson on its own — 「信任」 would
    # match a dozen.
    prefixed = [v for k, v in meta.items()
                if min(len(k), len(key)) >= 4 and (k.startswith(key) or key.startswith(k))]
    if len(prefixed) == 1:
        return prefixed[0]

    # Last resort, for titles that differ by a word rather than a subtitle
    # (「國高中數學當然可以…」 / 「國高中數學課當然可以…」). The threshold is high and
    # the match must be unambiguous; every pairing it makes is recorded in
    # metadata.yml, because a fuzzy join is how mis-bindings get made.
    import difflib

    near = difflib.get_close_matches(key, list(meta), n=2, cutoff=0.86)
    if len(near) == 1 or (len(near) == 2 and
                          difflib.SequenceMatcher(None, key, near[0]).ratio()
                          - difflib.SequenceMatcher(None, key, near[1]).ratio() > 0.06):
        return meta[near[0]]
    return None


def _headers(ws) -> dict[str, int]:
    """Header text → column index. Matched by prefix: the strategy column's full
    header runs past what fits and is truncated differently between sheets."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(c).strip(): i for i, c in enumerate(row) if c}


def _pick(headers: dict[str, int], *prefixes: str) -> int | None:
    for p in prefixes:
        for h, i in headers.items():
            if h.startswith(p):
                return i
    return None


def read_sheet(ws) -> dict[str, dict]:
    h = _headers(ws)
    col = {
        "title": _pick(h, "課名"),
        "genre": _pick(h, "文體"),
        "kind": _pick(h, "類型"),
        "topic": _pick(h, "單元議題"),
        "strategy": _pick(h, "閱讀聚光燈策略──教材目標策略", "閱讀聚光燈策略"),
        "heading": _pick(h, "閱讀聚光燈策略──第六大題標題"),
    }
    if col["title"] is None:
        return {}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[col["title"]] if col["title"] < len(row) else None
        if not title:
            continue
        get = lambda k: (str(row[col[k]]).strip()
                         if col[k] is not None and col[k] < len(row) and row[col[k]] else "")
        out.setdefault(norm_title(title), {
            "title": str(title).strip(),
            "genre": get("genre"),
            "kind": get("kind"),
            "unit_topic": get("topic"),
            "strategy": get("strategy"),
            "strategy_heading": get("heading"),
        })
    return out


def read_videos(ws) -> dict[str, list[str]]:
    h = _headers(ws)
    tcol = _pick(h, "課名")
    vcols = [i for name, i in h.items() if name.startswith("影片")]
    if tcol is None or not vcols:
        return {}
    out: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[tcol] if tcol < len(row) else None
        if not title:
            continue
        urls = [str(row[i]).strip() for i in vcols
                if i < len(row) and row[i] and str(row[i]).startswith("http")]
        if urls:
            out.setdefault(norm_title(title), urls)
    return out


def build_intro(meta: dict) -> str | None:
    """One sentence a student reads before starting.

    Built from the two columns that describe what the lesson is FOR — its unit topic
    and the reading strategy it teaches. Where neither exists there is no intro; a
    sentence assembled from the genre alone ("這是一篇記敘文") tells a reader nothing
    they cannot see from the card.
    """
    topic = meta.get("unit_topic", "")
    heading = meta.get("strategy_heading") or meta.get("strategy") or ""
    heading = re.sub(r"^.*?──\s*", "", heading).strip()
    if topic and heading:
        return f"本課圍繞「{topic}」，練習「{heading}」的閱讀方法。"
    if heading:
        return f"本課練習「{heading}」的閱讀方法。"
    if topic:
        return f"本課圍繞「{topic}」這個主題。"
    return None


def build(xlsx: Path) -> dict[str, dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    merged: dict[str, dict] = {}
    for name in wb.sheetnames:
        if name.startswith("4."):
            continue
        merged.update(read_sheet(wb[name]))
    videos = read_videos(wb[next(n for n in wb.sheetnames if n.startswith("4."))])

    for key, meta in merged.items():
        meta["intro"] = build_intro(meta)
        meta["category"] = GENRE_TO_CATEGORY.get(meta.get("genre", ""), "")
        meta["video_links"] = videos.get(key, [])
    return merged


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--out", default="/tmp/lesson_metadata.json")
    a = ap.parse_args()

    data = build(Path(a.xlsx))
    Path(a.out).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    have = lambda k: sum(1 for m in data.values() if m.get(k))
    print(f"  試算表 {len(data)} 課")
    print(f"    簡介 {have('intro')}   文體 {have('genre')}   分類 {have('category')}   影片 {have('video_links')}")
    sample = next(iter(data.values()))
    print(f"  範例《{sample['title']}》")
    print(f"    {sample['intro']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
