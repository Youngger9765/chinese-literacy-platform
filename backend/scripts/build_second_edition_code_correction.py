"""Build the 二修課號校正清單 for issue #2561 line ① (排序校正).

READ-ONLY. This script never writes to `backend/data/curriculum/manifest.yml`
or `backend/data/curriculum/lessons/*.yml`. It only reads the source-of-truth
xlsx (given by --xlsx, lives outside the repo) and the staging `/api/stories`
endpoint, and produces a review list (markdown + CSV) for Young / 淑麗 to sign
off before anyone touches the manifest.

Traps this script deliberately works around:

1. `1.總表!A` (舊課次) is mostly Excel-mangled dates (`m-d` numFmt). openpyxl's
   `data_only=False` load already returns a `datetime` for those cells (Excel
   applies the numFmt at write time), so we read `.month`/`.day` directly
   instead of re-deriving from a serial number.
2. The 二修 DOCX set only covers G4/G5/G8/G9 (73 課); G6/G7 (66 課) are
   untouched. This script does NOT filter by that — it processes all 159
   總表 rows plus the two side sheets, because the *ordering* correction is
   independent of whether the *content* has been rewritten yet.
3. `1.總表`(159) / `2.體育生的品格聚光燈`(16) / `3.文言文`(12) have three
   different column layouts. Each sheet gets its own parser.
4. Online `grade_code` has two incompatible formats co-existing in the same
   `stories` table: legacy low-`id` rows are zero-padded (`G4-L05`) and are
   frequently *stale duplicates* of a current high-`id` (1000+) unpadded row
   (`G4-L5`) — see `content-mapping-integrity` skill. This script builds a
   code index keyed on the *normalized* (grade, lesson_number, ab_suffix)
   tuple, so both formats collide into one bucket the way the real loader's
   `normalize_manifest_code` does, and reports the leftover duplicate id in
   the "線上多出" bucket with an explicit note instead of hiding it.
5. 課名 is not a unique key (「正太與小豬」appears once as a real G4 lesson
   and five more times as a `課次=0` 示範課 placeholder for other grades).
   Matching is primarily by the (grade, lesson) composite key reconstructed
   from 舊課次; title is only the tie-breaker when a code collides.

Usage:
    pip install openpyxl            # not in backend/requirements.txt — this
                                      # is an offline one-shot analysis tool,
                                      # not a deployed service dependency.
    python3 backend/scripts/build_second_edition_code_correction.py \\
        --xlsx "/path/to/自學教材總表.xlsx" \\
        --out-dir docs/curriculum
"""
from __future__ import annotations

import argparse
import csv
import datetime
import difflib
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.request import urlopen

STAGING_API = "https://lingoleap-backend-staging-958347263320.asia-east1.run.app"

# ---------------------------------------------------------------------------
# Multi-lesson secondary-slot overrides (issue #2561 finding).
#
# These (grade, lesson) keys are 舊課次 positions that were folded into a
# neighbouring lesson's combined DOCX/story a while ago — see
# `app/services/lesson_code_normalization.py` MULTI_LESSON_MAP for the
# authoritative Layer-2 file mapping this mirrors. There is no independent
# `stories` row for the secondary slot; the content lives inside the host
# story_id. Resolved empirically against the staging API (2026-08-14) rather
# than imported directly, because MULTI_LESSON_MAP maps *catalog codes* to
# *parsed YAML filenames*, not to `stories.id` — the two are a different axis
# and would silently mis-map if reused verbatim here.
# ---------------------------------------------------------------------------
SECONDARY_SLOT_HOST: dict[tuple[str, int], int] = {
    ("G4", 21): 1020,  # host: G4-L20 物以稀為貴 (供需多文本 2/3, 3/3)
    ("G4", 22): 1020,
    ("G5", 25): 1051,  # host: G5-L24 牧羊少年的逆轉勝 (多文本 2/2)
    ("G9", 16): 56,  # host: G9-L15 舊合併記錄(未解之謎 巨石陣+摩艾石像)；
    #                   高 id 1144 只拆出巨石陣一篇，摩艾石像仍只在舊合併記錄
    ("G9", 18): 58,  # host: G9-L17 舊合併記錄(馬拉松王者 3 篇)；
    ("G9", 19): 58,  # 高 id 1146 只拆出基普喬吉一篇，其餘兩篇仍只在舊合併記錄
}

# Reused-story-text rows: 總表「舊課次」是非日期字串(陷阱1 的 7 筆)，其中兩筆
# 是把既有課文重新包裝成另一個閱讀策略練習，而非獨立課文。沒有獨立 story_id，
# 只留références 供人工確認二修後是否仍要保留為獨立課。
REUSED_CONTENT_HINT: dict[str, int] = {
    "正太與小豬（自我提問策略-讀出關係）": 1007,  # 同 story 1007 正太與小豬
    "把手上的餅吃香一點（自我提問策略）": 1072,  # 同 story 1072/27 把手上的餅吃香一點
}


def normalize_title(s: Any) -> str:
    """Strip everything except Han characters/alnum so punctuation-only
    title drift (「」vs "", ？ vs ?, full-width vs half-width, IVS variation
    selectors) doesn't defeat matching."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return "".join(ch for ch in s if ("一" <= ch <= "鿿") or ch.isalnum())


CODE_RE = re.compile(r"^(G\d+|文)-L(\d+)([ab]?)$")


def parse_online_code(code: str) -> tuple[str, int, str] | None:
    m = CODE_RE.match(code)
    if not m:
        return None
    return (m.group(1), int(m.group(2)), m.group(3))


def grade_prefix(grade: int | str) -> str:
    return f"G{grade}"


# ---------------------------------------------------------------------------
# Online data
# ---------------------------------------------------------------------------


@dataclass
class OnlineLesson:
    id: int
    grade: str
    grade_code: str
    lesson_number: int
    title: str
    has_key_reading: bool
    norm_title: str = ""
    code_key: tuple[str, int, str] | None = None
    consumed_by: str | None = None  # sheet+row that claimed this id

    def __post_init__(self) -> None:
        self.norm_title = normalize_title(self.title)
        self.code_key = parse_online_code(self.grade_code)


def fetch_online_lessons(api_base: str) -> list[OnlineLesson]:
    url = f"{api_base.rstrip('/')}/api/stories?page_size=300"
    with urlopen(url, timeout=60) as resp:  # nosec B310: URL is operator-provided
        data = json.load(resp)
    stories = data["stories"]
    if len(stories) != data["total"]:
        raise SystemExit(
            f"story list truncated: got {len(stories)} / {data['total']} — "
            "raise page_size before trusting this script"
        )
    return [
        OnlineLesson(
            id=s["id"],
            grade=str(s["grade"]),
            grade_code=s["grade_code"],
            lesson_number=s["lesson_number"],
            title=s["title"],
            has_key_reading=bool(s.get("has_key_reading", False)),
        )
        for s in stories
    ]


# ---------------------------------------------------------------------------
# Excel parsing (three sheets, three layouts — trap 3)
# ---------------------------------------------------------------------------


@dataclass
class SourceRow:
    sheet: str
    excel_row: int
    title: str
    new_grade: Any
    new_lesson: Any
    new_code: str
    old_grade: int | None = None
    old_lesson: int | None = None
    old_raw: str | None = None
    old_source: str = "empty"  # date-serial / string / empty / string-unparsed
    genre: str | None = None
    is_special: str | None = None  # "demo_placeholder" / "authoring_placeholder" / None


def parse_main_sheet(ws) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(row=r, column=4).value
        if not title or not str(title).strip():
            continue
        a = ws.cell(row=r, column=1).value
        new_grade = ws.cell(row=r, column=2).value
        new_lesson = ws.cell(row=r, column=3).value
        genre = ws.cell(row=r, column=5).value

        old_grade = old_lesson = None
        old_raw = None
        old_source = "empty"
        if a is None or (isinstance(a, str) and not a.strip()):
            old_source = "empty"
        elif isinstance(a, (datetime.datetime, datetime.date)):
            old_grade, old_lesson = a.month, a.day
            old_raw = f"{old_grade}-{old_lesson}"
            old_source = "date-serial"
        elif isinstance(a, str):
            m = re.match(r"^(\d+)-(\d+)$", a.strip())
            if m:
                old_grade, old_lesson = int(m.group(1)), int(m.group(2))
                old_raw = a.strip()
                old_source = "string"
            else:
                old_raw = a.strip()
                old_source = "string-unparsed"

        is_special = "demo_placeholder" if new_lesson == 0 else None

        rows.append(
            SourceRow(
                sheet="1.總表",
                excel_row=r,
                title=str(title).strip(),
                new_grade=new_grade,
                new_lesson=new_lesson,
                new_code=f"G{new_grade}-L{int(new_lesson):02d}" if isinstance(new_lesson, int) else "",
                old_grade=old_grade,
                old_lesson=old_lesson,
                old_raw=old_raw,
                old_source=old_source,
                genre=genre,
                is_special=is_special,
            )
        )
    return rows


def parse_sports_sheet(ws) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if not code:
            continue
        orig_ref = ws.cell(row=r, column=2).value
        new_grade = ws.cell(row=r, column=3).value
        title = ws.cell(row=r, column=4).value
        title_s = str(title).strip() if title else ""
        is_special = "authoring_placeholder" if title_s == "待寫" else None
        rows.append(
            SourceRow(
                sheet="2.體育生的品格聚光燈",
                excel_row=r,
                title=title_s,
                new_grade=new_grade,
                new_lesson=None,
                new_code=str(code),
                old_raw=f"原課次(全域序號)={orig_ref}" if orig_ref else None,
                old_source="no-old-code-scheme",
                is_special=is_special,
            )
        )
    return rows


def parse_classical_sheet(ws) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        m = re.match(r"^文-(\d+)$", str(code).strip()) if code else None
        if not m:
            # Guards against stray leaked content from a neighbouring sheet
            # (row 16 here is a copy-paste artifact: "影片連結" table).
            continue
        idx = int(m.group(1))
        new_grade = ws.cell(row=r, column=2).value
        title = ws.cell(row=r, column=3).value
        genre = ws.cell(row=r, column=4).value
        title_s = str(title).strip() if title else ""
        is_special = "authoring_placeholder" if not title_s else None
        rows.append(
            SourceRow(
                sheet="3.文言文",
                excel_row=r,
                title=title_s,
                new_grade=new_grade,
                new_lesson=idx,
                new_code=f"文-L{idx:02d}" if idx else str(code),
                old_source="no-old-code-scheme",
                genre=genre,
                is_special=is_special,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------


@dataclass
class MatchResult:
    row: SourceRow
    online: OnlineLesson | None
    method: str  # code / code+tiebreak / title / override / reused-hint / none
    alt_candidates: list[OnlineLesson] = field(default_factory=list)
    note: str = ""


def build_online_indexes(
    online: list[OnlineLesson],
) -> tuple[dict[tuple[str, int, str], list[OnlineLesson]], dict[str, list[OnlineLesson]]]:
    by_code: dict[tuple[str, int, str], list[OnlineLesson]] = {}
    by_title: dict[str, list[OnlineLesson]] = {}
    for o in online:
        if o.code_key:
            by_code.setdefault(o.code_key, []).append(o)
        by_title.setdefault(o.norm_title, []).append(o)
    return by_code, by_title


def _score(cand: OnlineLesson, tnorm: str) -> tuple[int, int]:
    cn = cand.norm_title
    if cn == tnorm:
        s = 0
    elif tnorm and (tnorm in cn or cn in tnorm):
        s = 1
    else:
        s = 2
    return (s, -cand.id)  # prefer exact/substring match, then higher (current) id


def pick_best(candidates: list[OnlineLesson], title: str) -> OnlineLesson:
    tnorm = normalize_title(title)
    return sorted(candidates, key=lambda c: _score(c, tnorm))[0]


def _fuzzy_similar(a: str, b: str, threshold: float = 0.5) -> bool:
    """True if two normalized titles are close enough to be the same lesson.

    Guards the G8-mismatch rescue (below) against false positives from
    single-character text variants (e.g. 沒/没) that fail the strict
    exact/substring check in `_score` but are obviously the same title.
    """
    if not a or not b:
        return False
    return difflib.SequenceMatcher(None, a, b).ratio() >= threshold


def match_main_row(
    row: SourceRow,
    by_code: dict[tuple[str, int, str], list[OnlineLesson]],
    by_title: dict[str, list[OnlineLesson]],
    online_by_id: dict[int, OnlineLesson],
) -> MatchResult:
    if row.is_special == "demo_placeholder":
        tnorm = normalize_title(row.title)
        tcands = by_title.get(tnorm, [])
        chosen = pick_best(tcands, row.title) if tcands else None
        return MatchResult(row=row, online=chosen, method="demo-placeholder")

    if row.old_grade is not None:
        override_id = SECONDARY_SLOT_HOST.get((grade_prefix(row.old_grade), row.old_lesson))
        if override_id is not None:
            host = online_by_id[override_id]
            note = (
                f"舊課次 G{row.old_grade}-L{row.old_lesson} 是多文本合併課的副段落，"
                f"獨立內容併在 host story_id={override_id}（{host.grade_code} {host.title}），"
                "線上沒有屬於這個副段落自己的 story_id。"
            )
            return MatchResult(row=row, online=host, method="override-secondary-slot", note=note)

    reused_id = REUSED_CONTENT_HINT.get(row.title)
    if reused_id is not None:
        host = online_by_id[reused_id]
        note = (
            f"舊課次「{row.old_raw}」不是合法日期（陷阱1 的字串型別），推斷是把既有課文"
            f"「{host.title}」(story_id={reused_id}) 重新包裝成另一個閱讀策略練習，"
            "線上找不到屬於這個練習版本自己的 story_id，需人工確認二修後是否仍要保留為獨立課。"
        )
        return MatchResult(row=row, online=None, method="reused-content-hint", note=note)

    if row.old_grade is not None:
        key = (grade_prefix(row.old_grade), row.old_lesson, "")
        candidates = by_code.get(key, [])
        if candidates:
            chosen = pick_best(candidates, row.title) if len(candidates) > 1 else candidates[0]
            alts = [c for c in candidates if c is not chosen]
            method = "code" if len(candidates) == 1 else "code+title-tiebreak"

            # 陷阱4 的變形：即使代碼比對只有唯一候選，也不能無條件信任——
            # G8 因為 a/b 分裂課（G8-L3a/3b、G8-L6a/6b、G8-L9a/9b）造成往後的
            # 舊課次數字位移，唯一候選的課名可能其實是別的線上課。score==2 代表
            # 課名完全無關，這時改用課名比對搶救，搶救不到才承認比對失敗。
            tnorm = normalize_title(row.title)
            score = _score(chosen, tnorm)[0]
            if score == 2 and not _fuzzy_similar(chosen.norm_title, tnorm):
                tcands = by_title.get(tnorm, [])
                rescued = pick_best(tcands, row.title) if tcands else None
                mismatch_note = (
                    f"⚠️ 舊課次代碼比對到 story_id={chosen.id}（{chosen.grade_code} "
                    f"{chosen.title}），但課名跟總表列的「{row.title}」完全不同——"
                    "這是 G8 因 a/b 分裂課造成舊課次數字位移的已知現象，不是唯一候選"
                    "就代表比對正確。"
                )
                if rescued is not None:
                    mismatch_note += (
                        f"改用課名比對搶救為 story_id={rescued.id}（{rescued.grade_code}），"
                        "需人工確認此列的正確線上對應。"
                    )
                    return MatchResult(
                        row=row,
                        online=rescued,
                        method="code-title-mismatch-rescued-by-title",
                        alt_candidates=[chosen],
                        note=mismatch_note,
                    )
                mismatch_note += "課名比對也找不到線上對應，需人工核對，暫列總表有但線上無。"
                return MatchResult(
                    row=row,
                    online=None,
                    method="code-title-mismatch-unresolved",
                    alt_candidates=[chosen],
                    note=mismatch_note,
                )

            return MatchResult(row=row, online=chosen, method=method, alt_candidates=alts)

    tnorm = normalize_title(row.title)
    tcands = by_title.get(tnorm, [])
    if tcands:
        chosen = pick_best(tcands, row.title)
        alts = [c for c in tcands if c is not chosen]
        return MatchResult(row=row, online=chosen, method="title-fallback", alt_candidates=alts)

    return MatchResult(row=row, online=None, method="none")


def match_by_title_only(
    row: SourceRow, by_title: dict[str, list[OnlineLesson]]
) -> MatchResult:
    if row.is_special == "authoring_placeholder":
        return MatchResult(row=row, online=None, method="authoring-placeholder")
    tnorm = normalize_title(row.title)
    tcands = by_title.get(tnorm, [])
    if tcands:
        chosen = pick_best(tcands, row.title)
        alts = [c for c in tcands if c is not chosen]
        return MatchResult(row=row, online=chosen, method="title", alt_candidates=alts)
    return MatchResult(row=row, online=None, method="none")


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------


def classify(mr: MatchResult) -> str:
    row = mr.row
    if row.is_special == "demo_placeholder":
        return "不變"  # story_id reused as-is (§ demo placeholder note explains)
    if row.is_special == "authoring_placeholder":
        return "新增"
    if mr.online is None:
        if row.old_source == "empty" and row.sheet == "1.總表":
            return "新增"
        return "總表有但線上無"

    online = mr.online
    if online.code_key:
        online_grade = online.code_key[0]
    else:
        online_grade = None

    # secondary-slot / reused-content overrides already compare against the
    # row's own old_grade/old_lesson (recorded on the row), not the host's.
    old_grade_prefix = grade_prefix(row.old_grade) if row.old_grade is not None else online_grade
    new_grade_prefix = grade_prefix(row.new_grade) if isinstance(row.new_grade, int) else row.new_code[:2]

    if old_grade_prefix and new_grade_prefix and old_grade_prefix != new_grade_prefix:
        return "改code"
    if row.old_lesson is not None and row.new_lesson is not None and row.old_lesson != row.new_lesson:
        return "改序"
    if row.old_lesson is not None and row.new_lesson is not None and row.old_lesson == row.new_lesson:
        return "不變"
    return "改序"  # matched via title-fallback with no old_lesson to compare (blank 舊課次)


# ---------------------------------------------------------------------------
# Report rows
# ---------------------------------------------------------------------------


@dataclass
class ReportRow:
    sheet: str
    excel_row: int
    story_id: str
    online_grade_code: str
    online_lesson_number: str
    new_code: str
    old_code: str
    title: str
    label: str
    note: str


def build_report_rows(matches: list[MatchResult]) -> list[ReportRow]:
    out = []
    for mr in matches:
        row = mr.row
        label = classify(mr)
        old_code_display = (
            f"G{row.old_grade}-L{row.old_lesson}"
            if row.old_grade is not None
            else (row.old_raw or "（無，總表留白）")
        )
        note = mr.note
        if mr.alt_candidates:
            alt_desc = "; ".join(f"story_id={c.id}({c.grade_code} {c.title})" for c in mr.alt_candidates)
            dup_note = f"⚠️ 同 code/課名另有候選未採用（線上既有重複資料）：{alt_desc}"
            note = f"{note} {dup_note}".strip()
        if row.is_special == "demo_placeholder":
            note = (
                f"{note} 課次=0 是跨年級共用示範課占位（非獨立課次），"
                f"沿用 story_id={mr.online.id if mr.online else '?'}，不需要新分配課號。"
            ).strip()
        if row.is_special == "authoring_placeholder":
            note = f"{note} 總表標記「待寫」，教材尚未撰寫，非缺漏/非下架。".strip()

        out.append(
            ReportRow(
                sheet=row.sheet,
                excel_row=row.excel_row,
                story_id=str(mr.online.id) if mr.online else "",
                online_grade_code=mr.online.grade_code if mr.online else "",
                online_lesson_number=str(mr.online.lesson_number) if mr.online else "",
                new_code=row.new_code,
                old_code=old_code_display,
                title=row.title,
                label=label,
                note=note,
            )
        )
    return out


def build_orphan_rows(online: list[OnlineLesson], consumed_ids: set[int]) -> list[ReportRow]:
    """線上多出：online lessons no 總表/體育/文言 row ever claimed."""
    out = []
    dup_explained: dict[int, str] = {}
    # Pre-compute duplicate explanations: any online id sharing a code_key
    # with a consumed id is a legacy/current duplicate, not a real gap.
    by_code: dict[tuple[str, int, str], list[OnlineLesson]] = {}
    for o in online:
        if o.code_key:
            by_code.setdefault(o.code_key, []).append(o)
    for key, group in by_code.items():
        if len(group) < 2:
            continue
        consumed_in_group = [g for g in group if g.id in consumed_ids]
        unconsumed_in_group = [g for g in group if g.id not in consumed_ids]
        if consumed_in_group and unconsumed_in_group:
            for u in unconsumed_in_group:
                other = consumed_in_group[0]
                dup_explained[u.id] = (
                    f"與 story_id={other.id}（{other.grade_code} {other.title}）"
                    f"正規化後同碼 {key}，屬既有 DB 重複資料（補零/不補零兩筆），"
                    "已由另一筆對應二修總表，此筆非二修新增或下架。"
                )

    for o in online:
        if o.id in consumed_ids:
            continue
        note = dup_explained.get(
            o.id, "總表/體育/文言三個 sheet 都找不到對得上的 (年級,課次) 或課名，需人工確認是否為總表未列的既有教材。"
        )
        out.append(
            ReportRow(
                sheet="(線上)",
                excel_row=0,
                story_id=str(o.id),
                online_grade_code=o.grade_code,
                online_lesson_number=str(o.lesson_number),
                new_code="",
                old_code="",
                title=o.title,
                label="線上多出",
                note=note,
            )
        )
    return out


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

CSV_FIELDS = [
    "sheet",
    "excel_row",
    "story_id",
    "online_grade_code",
    "online_lesson_number",
    "new_code",
    "old_code",
    "title",
    "label",
    "note",
]


def write_csv(path: Path, rows: list[ReportRow]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "sheet": r.sheet,
                    "excel_row": r.excel_row,
                    "story_id": r.story_id,
                    "online_grade_code": r.online_grade_code,
                    "online_lesson_number": r.online_lesson_number,
                    "new_code": r.new_code,
                    "old_code": r.old_code,
                    "title": r.title,
                    "label": r.label,
                    "note": r.note,
                }
            )


def write_markdown(path: Path, rows: list[ReportRow], summary: dict[str, int]) -> None:
    lines = []
    lines.append("# 二修課號校正清單（issue #2561 ①排序校正）")
    lines.append("")
    lines.append(
        "⚠️ **這是唯讀分析結果，不是已執行的變更。** "
        "`backend/data/curriculum/manifest.yml` 與 `backend/data/curriculum/lessons/*.yml` "
        "未被本清單修改一個字。任何實際重新編號都要等 Young / 淑麗簽核後才進 manifest。"
    )
    lines.append("")
    lines.append(
        "⚠️ **story_id 契約**：QR 已印在紙本上，綁的是 `story_id`（不是 grade_code）。"
        "本清單完全不建議、也不涉及改動任何 story_id —— 見文末 "
        "`verify_qr_manifest.py` 的驗證結果。"
    )
    lines.append("")
    lines.append("## 摘要")
    lines.append("")
    lines.append("| 判定標籤 | 筆數 |")
    lines.append("|---|---|")
    for label, count in summary.items():
        lines.append(f"| {label} | {count} |")
    lines.append("")
    lines.append(
        "## 逐筆清單（依 sheet + excel_row 排序；線上多出附在各 sheet 之後）"
    )
    lines.append("")
    lines.append(
        "| sheet | excel_row | story_id | 線上grade_code | 線上lesson_number | 二修新課號 | 反解舊課次 | 課名 | 判定標籤 | 說明 |"
    )
    lines.append("|---|---|---|---|---|---|---|---|---|---|")
    for r in rows:
        note = r.note.replace("|", "\\|")
        title = r.title.replace("|", "\\|")
        lines.append(
            f"| {r.sheet} | {r.excel_row} | {r.story_id} | {r.online_grade_code} | "
            f"{r.online_lesson_number} | {r.new_code} | {r.old_code} | {title} | {r.label} | {note} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="path to 自學教材總表.xlsx (outside repo)")
    parser.add_argument("--api-base", default=STAGING_API)
    parser.add_argument("--out-dir", default="docs/curriculum")
    args = parser.parse_args(argv)

    import openpyxl  # local import: not a repo dependency, see module docstring

    wb = openpyxl.load_workbook(args.xlsx, data_only=False)
    main_rows = parse_main_sheet(wb["1.總表"])
    sports_rows = parse_sports_sheet(wb["2.體育生的品格聚光燈"])
    classical_rows = parse_classical_sheet(wb["3.文言文"])

    online = fetch_online_lessons(args.api_base)
    online_by_id = {o.id: o for o in online}
    by_code, by_title = build_online_indexes(online)

    matches: list[MatchResult] = []
    for row in main_rows:
        mr = match_main_row(row, by_code, by_title, online_by_id)
        matches.append(mr)
    for row in sports_rows:
        matches.append(match_by_title_only(row, by_title))
    for row in classical_rows:
        matches.append(match_by_title_only(row, by_title))

    consumed_ids = {mr.online.id for mr in matches if mr.online is not None}

    report_rows = build_report_rows(matches)
    orphan_rows = build_orphan_rows(online, consumed_ids)
    all_rows = report_rows + orphan_rows

    summary: dict[str, int] = {}
    for r in all_rows:
        summary[r.label] = summary.get(r.label, 0) + 1

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "2026-08-14-second-edition-lesson-code-correction.csv"
    md_path = out_dir / "2026-08-14-second-edition-lesson-code-correction.md"
    write_csv(csv_path, all_rows)
    write_markdown(md_path, all_rows, summary)

    print(f"wrote {csv_path}")
    print(f"wrote {md_path}")
    print("summary:", summary)
    print(f"total rows: {len(all_rows)} (main={len(main_rows)} sports={len(sports_rows)} "
          f"classical={len(classical_rows)} orphan={len(orphan_rows)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
