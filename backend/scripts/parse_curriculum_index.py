#!/usr/bin/env python3
"""
parse_curriculum_index.py
解析陳教授 Excel 策略總表 (自學教材清單.xlsx) → curriculum-index.json

設計原則：Excel 所有欄位 1:1 保留，column name 直接當 JSON key。
不過濾、不重新命名、不重新組織。

Usage:
    python3 backend/scripts/parse_curriculum_index.py \
        --xlsx "private/curriculum-source/2026-05-01/1.L1-158新版完成學習單1150415/自學教材清單.xlsx" \
        --output backend/data/curriculum-index.json \
        --lessons-dir backend/data/lessons

Part of issue #1348 (sub-task of #1344)
"""

import argparse
import datetime
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

EXCEL_DATE_BASE = datetime.datetime(2025, 4, 1)


def decode_xin_keci(value):
    """
    新課次 in 總表 and 影片連結-新 is stored as an Excel date cell (format: m-d).
    E.g. 2025-04-01 → display '4-1' → decoded as new_course_number = 1.
    Returns (display_str, integer) tuple.
    """
    if value is None:
        return None, None
    if isinstance(value, datetime.datetime):
        display = f"{value.month}-{value.day}"
        number = (value - EXCEL_DATE_BASE).days + 1
        return display, number
    # 注意課次安排 and 文言文 store as plain integer
    try:
        n = int(value)
        return str(n), n
    except (ValueError, TypeError):
        return str(value), None


def cell_value(row, idx, headers=None):
    """
    Safely get cell value at index idx.
    Returns None if out-of-bounds.
    Normalises datetime → ISO string to preserve raw value.
    Normalises float that is whole number → int (e.g. 4.0 → 4).
    """
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        # Keep as ISO string — caller handles display conversion
        return v.isoformat()
    if isinstance(v, float):
        if v != v:  # NaN check
            return None
        if v == int(v):
            return int(v)
        return v
    if isinstance(v, str):
        stripped = v.strip()
        if stripped == "" or stripped == "\n":
            return None
        return stripped
    return v


def row_to_dict(row, headers):
    """
    Convert a row tuple to dict using headers as keys.
    Headers with None (unnamed columns) are included as 'unnamed_col_N'.
    """
    result = {}
    for i, h in enumerate(headers):
        key = h if h is not None else f"unnamed_col_{i}"
        result[key] = cell_value(row, i)
    return result


def load_yaml_titles(lessons_dir: str) -> dict:
    """Return {filename: title} for all .yml files in lessons_dir."""
    titles = {}
    if not os.path.isdir(lessons_dir):
        return titles
    for fname in sorted(os.listdir(lessons_dir)):
        if not fname.endswith(".yml"):
            continue
        fpath = os.path.join(lessons_dir, fname)
        try:
            with open(fpath, encoding="utf-8") as f:
                for line in f:
                    if line.startswith("title:"):
                        raw = line.split(":", 1)[1].strip().strip("'\"")
                        titles[fname] = raw
                        break
        except OSError:
            pass
    return titles


def norm_title(s: str) -> str:
    """Normalize title for fuzzy matching."""
    s = s.replace("──", "").replace("─", "").replace("—", "").replace("－", "")
    s = s.replace("没", "沒")
    s = re.sub(r"[，。、：；？！「」『』【】（）《》\-\s_—–《》＃#@]", "", s)
    return s.lower()


def match_yaml(title: str, yaml_titles: dict) -> str | None:
    """Find matching YAML file for a lesson title."""
    if not title:
        return None
    # Exact
    for fname, yt in yaml_titles.items():
        if yt == title:
            return fname
    # Normalized exact
    nt = norm_title(title)
    for fname, yt in yaml_titles.items():
        if norm_title(yt) == nt:
            return fname
    # Substring
    for fname, yt in yaml_titles.items():
        ny = norm_title(yt)
        if nt in ny or ny in nt:
            return fname
    return None


# ---------------------------------------------------------------------------
# Sheet parsers — all return list of dicts with raw column names as keys
# ---------------------------------------------------------------------------

def parse_sheet_raw(ws, sheet_name: str) -> tuple[list[str], list[dict]]:
    """
    Generic parser: reads all rows from a sheet.
    Returns (headers, rows) where headers are the column names from row 1,
    and rows are list of dicts with all column values.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        return [], []

    # Use column name as key; None → 'unnamed_col_N'
    headers = []
    for i, h in enumerate(header_row):
        if h is not None and str(h).strip():
            headers.append(str(h).strip())
        else:
            headers.append(None)  # will become unnamed_col_N in row_to_dict

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row_to_dict(row, headers)
        # Skip completely empty rows
        if all(v is None for v in d.values()):
            continue
        rows.append(d)

    return headers, rows


def parse_zongbiao(ws, yaml_titles: dict) -> list[dict]:
    """
    Parse 總表 sheet with all columns 1:1.
    Adds two computed fields:
      _新課次_display : '4-1', '4-2', ...  (month-day from Excel date)
      _新課次_number  : 1, 2, 3, ...       (sequential lesson number)
      _existing_yaml : matched YAML filename or null
    """
    headers, raw_rows = parse_sheet_raw(ws, "總表")

    lessons = []
    for d in raw_rows:
        # Only include rows with 課名
        if d.get("課名") is None:
            continue

        # Decode 新課次 Excel date → display + number
        xk_raw = d.get("新課次")
        if xk_raw and isinstance(xk_raw, str) and "T" in xk_raw:
            # ISO datetime string from cell_value conversion
            dt = datetime.datetime.fromisoformat(xk_raw)
            display = f"{dt.month}-{dt.day}"
            number = (dt - EXCEL_DATE_BASE).days + 1
        elif xk_raw is not None:
            try:
                number = int(xk_raw)
                display = str(number)
            except (ValueError, TypeError):
                display = str(xk_raw)
                number = None
        else:
            display, number = None, None

        # Replace the raw ISO datetime with the human-readable display value
        # but also add decoded integer
        d["新課次"] = display  # preserve as-displayed: '4-1', '4-2'...
        d["_新課次_number"] = number  # decoded integer 1, 2, 3...
        d["_source_sheet"] = "總表"
        d["_existing_yaml"] = match_yaml(d.get("課名", ""), yaml_titles)

        lessons.append(d)

    return lessons


def parse_wenyanwen(ws, yaml_titles: dict) -> list[dict]:
    """
    Parse 文言文 sheet with all columns 1:1.
    新課次 in this sheet is already an integer (169, 170...).
    Skips rows where 課名 looks like a URL or 新課次 is not an integer.
    """
    headers, raw_rows = parse_sheet_raw(ws, "文言文")

    lessons = []
    for d in raw_rows:
        title = d.get("課名")
        if title is None:
            continue
        # Skip rows where 課名 is a URL (junk row at bottom of sheet)
        if isinstance(title, str) and title.startswith("http"):
            continue
        # Skip rows where 新課次 is not an integer (e.g. a title string leaked in)
        xk = d.get("新課次")
        if not isinstance(xk, int):
            continue
        # 新課次 is integer in 文言文 sheet
        d["_source_sheet"] = "文言文"
        d["_新課次_number"] = xk  # already int
        d["_existing_yaml"] = match_yaml(title, yaml_titles)
        lessons.append(d)

    return lessons


def parse_zhuyi_keci(ws) -> list[dict]:
    """Parse 注意課次安排 sheet with all columns 1:1."""
    headers, raw_rows = parse_sheet_raw(ws, "注意課次安排")
    result = []
    for d in raw_rows:
        if d.get("課名") is None and d.get("新課次") is None:
            continue
        d["_source_sheet"] = "注意課次安排"
        result.append(d)
    return result


def parse_yixie_renwu(ws) -> list[dict]:
    """Parse 已寫過的人物 sheet with all columns 1:1."""
    headers, raw_rows = parse_sheet_raw(ws, "已寫過的人物")
    result = []
    for d in raw_rows:
        # Keep only rows with at least one non-null value
        if any(v is not None for v in d.values()):
            d["_source_sheet"] = "已寫過的人物"
            result.append(d)
    return result


def parse_yingpian(ws) -> list[dict]:
    """
    Parse 影片連結-新 sheet with all columns 1:1.
    新課次 here is also an Excel date cell.
    """
    headers, raw_rows = parse_sheet_raw(ws, "影片連結-新")

    result = []
    for d in raw_rows:
        if d.get("課名") is None and d.get("新課次") is None:
            continue
        # Decode 新課次
        xk_raw = d.get("新課次")
        if xk_raw and isinstance(xk_raw, str) and "T" in xk_raw:
            dt = datetime.datetime.fromisoformat(xk_raw)
            display = f"{dt.month}-{dt.day}"
            number = (dt - EXCEL_DATE_BASE).days + 1
        elif xk_raw is not None:
            try:
                number = int(xk_raw)
                display = str(number)
            except (ValueError, TypeError):
                display = str(xk_raw)
                number = None
        else:
            display, number = None, None

        d["新課次"] = display
        d["_新課次_number"] = number
        d["_source_sheet"] = "影片連結-新"
        result.append(d)

    return result


# ---------------------------------------------------------------------------
# Audit helpers
# ---------------------------------------------------------------------------

def build_audit(
    zongbiao: list[dict],
    wenyanwen: list[dict],
    zhuyi: list[dict],
    yaml_titles: dict,
) -> dict:
    """Build audit statistics dict."""
    from collections import Counter

    all_main = zongbiao + wenyanwen
    total = len(all_main)

    grade_dist = Counter(
        str(d.get("適用年級", "?")) for d in all_main
        if d.get("適用年級") is not None
    )

    strategy_col = "閱讀聚光燈策略──教材目標策略（學習單第一頁右上方）"
    strategy_dist = Counter(
        d.get(strategy_col) for d in zongbiao
        if d.get(strategy_col) is not None
    )

    # Existing YAML matches
    matched = [d for d in all_main if d.get("_existing_yaml")]
    unmatched_titles = [
        d.get("課名") for d in all_main if not d.get("_existing_yaml")
    ]

    # YAML files not covered by Excel
    matched_yamls = {d["_existing_yaml"] for d in matched if d["_existing_yaml"]}
    uncovered_yaml = [f for f in yaml_titles if f not in matched_yamls]

    # Grade 適用年級 coverage in 注意課次安排 vs 總表
    zhuyi_titles = {d.get("課名") for d in zhuyi if d.get("課名")}

    return {
        "total_main_lessons": len(zongbiao),
        "total_wenyanwen_lessons": len(wenyanwen),
        "total_all": total,
        "total_existing_yaml": len(yaml_titles),
        "matched_to_yaml": len(matched),
        "unmatched_excel_titles": unmatched_titles,
        "yaml_files_not_in_excel": uncovered_yaml,
        "grade_distribution": dict(sorted(grade_dist.items())),
        "strategy_distribution_top20": dict(strategy_dist.most_common(20)),
        "注意課次安排_lesson_count": len(zhuyi),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Parse curriculum Excel → curriculum-index.json (all columns 1:1)"
    )
    parser.add_argument(
        "--xlsx",
        default="private/curriculum-source/2026-05-01/1.L1-158新版完成學習單1150415/自學教材清單.xlsx",
    )
    parser.add_argument(
        "--output",
        default="backend/data/curriculum-index.json",
    )
    parser.add_argument(
        "--lessons-dir",
        default="backend/data/lessons",
    )
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: Excel file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    yaml_titles = load_yaml_titles(args.lessons_dir)
    print(f"Existing YAML files: {len(yaml_titles)}")

    # Parse all 5 sheets
    zongbiao     = parse_zongbiao(wb["總表"], yaml_titles)
    wenyanwen    = parse_wenyanwen(wb["文言文"], yaml_titles)
    zhuyi        = parse_zhuyi_keci(wb["注意課次安排"])
    yixie        = parse_yixie_renwu(wb["已寫過的人物"])
    yingpian     = parse_yingpian(wb["影片連結-新"])

    print(f"  總表:      {len(zongbiao)} lessons")
    print(f"  文言文:    {len(wenyanwen)} lessons")
    print(f"  注意課次安排: {len(zhuyi)} rows")
    print(f"  已寫過的人物: {len(yixie)} rows")
    print(f"  影片連結-新:  {len(yingpian)} rows")

    audit = build_audit(zongbiao, wenyanwen, zhuyi, yaml_titles)

    output = {
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "source_file": os.path.basename(args.xlsx),
        "source_path": args.xlsx,
        "sheets": {
            "總表": {
                "columns": [
                    "新課次", "原課次", "適用年級", "課名", "文體", "類型",
                    "閱讀聚光燈策略──教材目標策略（學習單第一頁右上方）",
                    "閱讀聚光燈策略──研發夥伴用", "單元議題",
                    "閱讀聚光燈策略──第六大題標題", "文本類型",
                    "語詞1", "語詞2", "語詞3", "語詞4", "語詞5", "語詞6",
                    "語詞7", "語詞8", "語詞9", "語詞10", "語詞11", "語詞12",
                    "語詞13", "語詞14", "語詞15", "語詞16", "語詞17", "語詞18", "語詞19",
                ],
                "row_count": len(zongbiao),
                "rows": zongbiao,
            },
            "文言文": {
                "columns": [
                    "新課次", "原課次", "適用年級", "課名", "文體", "類型", "單元議題",
                    "閱讀聚光燈策略──研發夥伴用",
                    "閱讀聚光燈策略──教材目標策略（學習單第一頁右上方）",
                    "閱讀聚光燈策略──第六大題標題", "文本類型",
                    "語詞1", "語詞2", "語詞3", "語詞4", "語詞5", "語詞6",
                    "語詞7", "語詞8", "語詞9", "語詞10", "語詞11", "語詞12",
                    "語詞13", "語詞14", "語詞15", "語詞16", "語詞17", "語詞18",
                    "語詞19", "語詞20", "負責人",
                ],
                "row_count": len(wenyanwen),
                "rows": wenyanwen,
            },
            "注意課次安排": {
                "columns": [
                    "新課次", "原課次", "文本類型", "課名", "文體", "單元議題",
                    "適用年級", "閱讀聚光燈策略", "因策略", "策略關聯",
                ],
                "row_count": len(zhuyi),
                "rows": zhuyi,
            },
            "已寫過的人物": {
                "columns": ["(unnamed)", "運動類", "古人", "現代人"],
                "row_count": len(yixie),
                "rows": yixie,
            },
            "影片連結-新": {
                "columns": ["新課次", "原課次", "課名", "影片1", "影片2", "影片3"],
                "row_count": len(yingpian),
                "rows": yingpian,
            },
        },
        "audit": audit,
    }

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nWritten: {args.output}")

    # Print audit summary
    print("\n=== Audit Summary ===")
    print(f"總表 lessons:          {audit['total_main_lessons']}")
    print(f"文言文 lessons:        {audit['total_wenyanwen_lessons']}")
    print(f"Total:                 {audit['total_all']}")
    print(f"Existing YAML:         {audit['total_existing_yaml']}")
    print(f"Matched to YAML:       {audit['matched_to_yaml']}")
    print(f"Grade distribution:    {audit['grade_distribution']}")
    print(f"Strategy count:        {len(audit['strategy_distribution_top20'])}")
    print(f"\nUnmatched Excel titles ({len(audit['unmatched_excel_titles'])}):")
    for t in audit["unmatched_excel_titles"][:20]:
        print(f"  - {t}")
    if len(audit["unmatched_excel_titles"]) > 20:
        print(f"  ... and {len(audit['unmatched_excel_titles']) - 20} more")
    print(f"\nYAML files not in Excel ({len(audit['yaml_files_not_in_excel'])}):")
    for f in audit["yaml_files_not_in_excel"]:
        print(f"  - {f}")


if __name__ == "__main__":
    main()
